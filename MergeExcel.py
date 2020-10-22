import pandas as pd
from openpyxl import load_workbook
import win32com.client as win32

excel1 = 'Fiberglass_ISF_22-Jan-2020.xlsx'
excel2 = 'Folding Partitions_ISF_22-Jan-2020.xlsx'
excel3 = 'Louvers_ISF_22-Jan-2020.xlsx'
excel4 = 'Polycarbonate_ISF_22-Jan-2020.xlsx'
excel5 = 'Skylights_ISF_22-Jan-2020.xlsx'
excel6 = 'Tubes_ISF_22-Jan-2020.xlsx'
exceljoin = 'ISFsearches.xlsx'
df1 = pd.read_excel(excel1)
df2 = pd.read_excel(excel2)
df3 = pd.read_excel(excel3)
df4 = pd.read_excel(excel4)
df5 = pd.read_excel(excel5)
df6 = pd.read_excel(excel6)
values1 = df1[['Project ID','Bid Date', 'Project Name', 'Posted By','City','Addenda Count',
'Last Addenda', 'Owner','Start Date','Project Url','Product Indicator']]
values2 = df2[['Project ID','Bid Date', 'Project Name', 'Posted By','City','Addenda Count',
'Last Addenda', 'Owner','Start Date','Project Url','Product Indicator']]
values3 = df3[['Project ID','Bid Date', 'Project Name', 'Posted By','City','Addenda Count',
'Last Addenda', 'Owner','Start Date','Project Url','Product Indicator']]
values4 = df4[['Project ID','Bid Date', 'Project Name', 'Posted By','City','Addenda Count',
'Last Addenda', 'Owner','Start Date','Project Url','Product Indicator']]
values5 = df5[['Project ID','Bid Date', 'Project Name', 'Posted By','City','Addenda Count',
'Last Addenda', 'Owner','Start Date','Project Url','Product Indicator']]
values6 = df6[['Project ID','Bid Date', 'Project Name', 'Posted By','City','Addenda Count',
'Last Addenda', 'Owner','Start Date','Project Url','Product Indicator']]

dataframes = [values1, values2, values3, values4, values5,values6]
join = pd.concat(dataframes)
join.sort_values('Project Name').to_excel('ISFsearches.xlsx', index=False)

dfjoin = pd.read_excel(exceljoin)

dfjoin[~dfjoin.duplicated(subset=['Project Name'])].sort_values('Project Name').to_excel('unique.xlsx', index=False)
dfjoin[dfjoin.duplicated(subset=['Project Name'])].sort_values('Project Name').to_excel('dupes.xlsx', index=False)

wsuni = load_workbook('unique.xlsx')
wsdupe = load_workbook('dupes.xlsx')
uni = wsuni['Sheet1']
dupe = wsdupe['Sheet1']

dupemaxr = dupe.max_row - 1
unimaxr = uni.max_row + 1
rows = 1
unirows = 1
pnamecol = 3
prodcol = 11

origlastcol = uni.max_column 
for row in dupe.iter_rows(max_col=1, max_row=dupemaxr):
	rows = rows + 1
	dupesName = (dupe.cell(row=rows, column=pnamecol).value)
	dupesProd = (dupe.cell(row=rows, column=prodcol).value)
	for row in uni.iter_rows(max_col=1, max_row=unimaxr):
		unirows = unirows + 1
		uniName = (uni.cell(row=unirows, column=pnamecol).value)
		uniProd = (uni.cell(row=unirows, column=origlastcol).value)
		if dupesName == uniName:
			lastcol = origlastcol
			uni.cell(row=unirows, column=origlastcol).value = str(uniProd) + ', ' + str(dupesProd)
	unirows = 1
wsuni.save('condensed.xlsx')